import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)

#wb.active selects the first avail sheet (sheet1)

sheet = wb.active
print(sheet.active)

for i in range(1, sheet.max_row+1):
    #print all data in column A,B,C
    #ljust() is used to make it easier to read
    print(i, str(sheet.cell(row = i, column = 1). value).ljust(25, '')),
    sheet.cell(row = i, column = 2). value.ljust(15, ''),
    str(sheet.cell(row = i, column = 3).value)